from openpyxl import Workbook

from permanent_output import permanent_output_func


def write_exel_table_func(name_file, response, oneall):
    if permanent_output_func("exel_output_file\\permanent_output_exel.txt") == "True":
        if oneall == 'one':
            wb = Workbook()
            ws = wb.active

            ws['A1'] = response[1]
            ws['A2'] = response[2]
            ws['A3'] = response[3]
            ws['A4'] = response[4]
            ws['A5'] = response[5]
            ws['A6'] = response[6]
            ws['A7'] = response[7]
            ws['A8'] = response[8]
            ws['A9'] = response[9]
            ws['A10'] = response[10]
            ws['A11'] = response[11]
            ws['A12'] = response[12]
            ws['A13'] = response[13]
            ws['A14'] = response[14]
            ws['A15'] = response[15]
            ws['A16'] = response[16]
            ws['A17'] = response[17]
            ws['A18'] = response[18]
            ws['A19'] = response[19]
            ws['A20'] = response[20]
            ws['A21'] = response[21]
            ws['A22'] = response[22]
            ws['A23'] = response[23]
            wb.save(f"exel_output_file\\exel_output_file_one\\{name_file}.xlsx")
        elif oneall == 'all':
            pass
    elif permanent_output_func("exel_output_file\\permanent_output_exel.txt") == "False":
        pass
